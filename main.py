from openpyxl import Workbook
import codecs
import pandas as pd
from datetime import datetime

wb = Workbook()
ws = wb.active

def geraDictFornecedores():
        fornecedores = dict()

        # read_file = pd.read_excel ('csv\\Fornecedor.xlsx')
        # read_file.to_csv ('csv\\Fornecedor.csv', index = None, header=True,sep="ϡ")

        with codecs.open('csv\\Fornecedor.csv',encoding='utf-8') as csv_file_fornec:
                next(csv_file_fornec, None)
                for line in csv_file_fornec:
                        x =  line.strip().split('ϡ')
                        nome_fornec = x[2]
                        numero_fornec_oracle = x[0]
                        cgc = x[1]

                        fornecedores[cgc] = [cgc, numero_fornec_oracle, nome_fornec]
                return fornecedores
fornecedores =  geraDictFornecedores()

def geraDictEmitentes():
    emitentes = dict()

#     read_file = pd.read_excel ('csv\\emitentes.xlsx')
#     read_file.to_csv ('csv\\emitentes.csv', index = None, header=True,sep="ϡ")

    with codecs.open('csv\\emitentes.csv',encoding='utf-8') as csv_file_emitente:
        # next(csv_file_emitente, None)
        for line in csv_file_emitente:
            x =  line.strip().replace('"','').split('ϡ')
            if len(x) == 4:
                # print(x)
            # print(x)
                cod_emitente = x[0]
                nome_emitente = x[1]
                cgc = x[2]
            
                emitentes[cod_emitente] = [cod_emitente, nome_emitente, cgc]
        return emitentes
emitentes =  geraDictEmitentes()

def geraHeader():
        dictHeader = dict()
        with codecs.open('csv\\tit_ap_adiantamento.csv',encoding='utf-8') as tit_ap:
                next(tit_ap, None)
                idHeader = 0
                for line in tit_ap:
                        idHeader += 1
                        formatedLine =  line.strip().split("ϡ")
                        # print(formatedLine)
                        fornecedor = formatedLine[6]

                        try:
                                formatedLine.append(emitentes[fornecedor][1])
                        except:
                                continue
                        # esse eu peguei direto do emitentes
                        cgc_emitentes = emitentes[fornecedor][2]
                        cod_emitentes = emitentes[fornecedor][0]
                        try:
            
                                cgc_fornecedores = fornecedores[cgc_emitentes][0]
                        except:
                                try:
                                        cgc_fornecedores = fornecedores['0'+cgc_emitentes][0]
                                        cgc_emitentes = '0'+cgc_emitentes
                                except:
                                        try:
                                                cgc_fornecedores = fornecedores[cgc_emitentes[1:]][0]
                                                cgc_emitentes = cgc_emitentes[1:]
                                        except:
                                                print(cgc_emitentes)
                                                continue
                                        
                        if fornecedor in emitentes:               
                                numNFF = formatedLine[1]
                                saldo = formatedLine[2].replace('.', ',')
                                dataNF = formatedLine[3]
                                cnpj = formatedLine[6]
                                dataVencimento = formatedLine[10]
                                nomeFornecedor = emitentes[fornecedor][1]
                                dictHeader[idHeader,numNFF,cnpj] = [idHeader,numNFF,saldo, dataNF, nomeFornecedor, cgc_fornecedores, dataVencimento]
                return dictHeader
dictHeader = geraHeader()
                    

def geraLine():
        dictLine = dict()
        with codecs.open('csv\\tit_ap_adiantamento.csv',encoding='utf-8') as tit_ap_line:
                next(tit_ap_line, None)
                idHeader = 0
                for line in tit_ap_line:
                        idHeader += 1
                        formatedLine =  line.strip().split("ϡ")
                        fornecedor = formatedLine[6]
                        try:
                                formatedLine.append(emitentes[fornecedor][1])
                        except:
                                continue
                        # esse eu peguei direto do emitentes
                        cgc_emitentes = emitentes[fornecedor][2]
                        cod_emitentes = emitentes[fornecedor][0]
                        try:
            
                                cgc_fornecedores = fornecedores[cgc_emitentes][0]
                        except:
                                try:
                                        cgc_fornecedores = fornecedores['0'+cgc_emitentes][0]
                                        cgc_emitentes = '0'+cgc_emitentes
                                except:
                                        try:
                                                cgc_fornecedores = fornecedores[cgc_emitentes[1:]][0]
                                                cgc_emitentes = cgc_emitentes[1:]
                                        except:
                                                # print(cgc_emitentes)
                                                continue
                        if fornecedor in emitentes:
                                numNFF = formatedLine[1]
                                saldo = formatedLine[2].replace('.', ',')
                                dataNF = formatedLine[3]
                                cnpj = formatedLine[6]
                                dataVencimento = formatedLine[10]
                                dictLine[idHeader,numNFF,cnpj] = [idHeader,numNFF,saldo, dataNF,fornecedor, cnpj, dataVencimento]
                return dictLine
dictLine = geraLine()

def header():
        wb = Workbook()
        ws = wb.active
        for line in dictHeader:
                row = []   
                idHeader = dictHeader[line][0]
                numNFF = dictHeader[line][1]
                saldo = dictHeader[line][2]

                # TRATATIVA DA DATA NF
                dt_nff = dictHeader[line][3].replace('-','/')
                DATA_NFF_DATE = datetime.strptime(dt_nff,'%Y/%m/%d').date()
                dataNF =  DATA_NFF_DATE.strftime('%d/%m/%Y')
                # 

                fornecedor = dictHeader[line][4]
                cnpj = dictHeader[line][5]
                
                # TRATATIVA DA DATA DE VENCIMENTO
                dt_venc = dictHeader[line][6].replace('-','/')
                DATA_VENCIMENTO_DATE = datetime.strptime(dt_venc,'%Y/%m/%d').date()
                dataVencimento = DATA_VENCIMENTO_DATE.strftime('%d/%m/%Y')
                # 
                

                row.append(idHeader)#*Invoice ID
                row.append("CEDISA_BU")#*Business Unit
                row.append("CARGA")#*Source
                row.append("ADTO_"+numNFF)#*Invoice Number
                row.append(saldo)# *Invoice Amount
                row.append(dataNF)# *Invoice Date
                row.append(fornecedor)# **Supplier Name
                row.append("")# **Supplier Number
                row.append(cnpj)# *Supplier Site
                row.append("BRL")# Invoice Currency 
                row.append("BRL")# Payment Currency
                row.append("CARGA DE SALDOS - ADIANTAMENTOS")# Description
                row.append("")# Import Set
                row.append("PREPAYMENT")# *Invoice Type
                row.append("")# Legal Entity
                row.append("")# Customer Tax Registration Number
                row.append("")# Customer Registration Code
                row.append("")# First-Party Tax Registration Number
                row.append("")# Supplier Tax Registration Number
                row.append("IMEDIATO")# *Payment Terms
                row.append(dataVencimento)# Terms Date
                row.append("")# Goods Received Date
                row.append("")# Invoice Received Date
                row.append("29/03/2022")# Accounting Date
                row.append("BR_BAIXA_MANUAL")# Payment Method
                row.append("ADIANTAMENTO")# Pay Group
                row.append("N")# Pay Alone
                row.append("")# Discountable Amount
                row.append("")# Prepayment Number
                row.append("")# Prepayment Line Number
                row.append("")# Prepayment Application Amount
                row.append("")# Prepayment Accounting Date
                row.append("")# Invoice Includes Prepayment
                row.append("")# Conversion Rate Type
                row.append("")# Conversion Date
                row.append("")# Conversion Rate
                row.append("1001.000.000.21110001.0000.0000.0.0")# Liability Combination
                row.append("")# Document Category Code
                row.append("")# Voucher Number
                row.append("")# Requester First Name
                row.append("")# Requester Last Name
                row.append("")# Requester Employee Number
                row.append("")# Delivery Channel Code
                row.append("")# Bank Charge Bearer
                row.append("")# Remit-to Supplier
                row.append("")# Remit-to Supplier Number
                row.append("")# Remit-to Address Name
                row.append("1")# Payment Priority
                row.append("")# Settlement Priority
                row.append("")# Unique Remittance Identifier
                row.append("")# Unique Remittance Identifier Check Digit
                row.append("")# Payment Reason Code
                row.append("")# Payment Reason Comments
                row.append("")# Remittance Message 1
                row.append("")# Remittance Message 2
                row.append("")# Remittance Message 3
                row.append("")# Withholding Tax Group
                row.append("")# Ship-to Location
                row.append("")# Taxation Country
                row.append("")# Document Sub Type
                row.append("")# Tax Invoice Internal Sequence Number
                row.append("")# Supplier Tax Invoice Number
                row.append("")# Tax Invoice Recording Date
                row.append("")# Supplier Tax Invoice Date
                row.append("")# Supplier Tax Invoice Conversion Rate
                row.append("")# Port Of Entry Code
                row.append("")# Correction Year
                row.append("")# Correction Period
                row.append("")# Import Document Number
                row.append("")# Import Document Date
                row.append("")# Tax Control Amount
                row.append("")# Calculate Tax During Import
                row.append("")# Add Tax To Invoice Amount
                row.append("")# Attribute Category
                row.append("")# Attribute 1
                row.append("")# Attribute 2
                row.append("")# Attribute 3
                row.append("")# Attribute 4
                row.append("")# Attribute 5
                row.append("")# Attribute 6
                row.append("")# Attribute 7
                row.append("")# Attribute 8
                row.append("")# Attribute 9
                row.append("")# Attribute 10
                row.append("")# Attribute 11
                row.append("")# Attribute 12
                row.append("")# Attribute 13
                row.append("")# Attribute 14
                row.append("")# Attribute 15
                row.append("")# Attribute Number 1
                row.append("")# Attribute Number 2
                row.append("")# Attribute Number 3
                row.append("")# Attribute Number 4
                row.append("")# Attribute Number 5
                row.append("")# Attribute Date 1
                row.append("")# Attribute Date 2
                row.append("")# Attribute Date 3
                row.append("")# Attribute Date 4
                row.append("")# Attribute Date 5
                row.append("JL_BR_APXINWKB_AP_INVOICES")# Global Attribute Category
                row.append("N")# Global Attribute 1
                
                ws.append(row)
        wb.save("Plan_Contas_AP_ADIANTAMENTO_HEADER_TESTE.xlsx")

header()

def line():
        wb = Workbook()
        ws = wb.active
        for line in dictLine:
                row = []   
                idHeader = dictLine[line][0]
                numNFF = dictLine[line][1]
                saldo = dictLine[line][2]
                dataNF = dictLine[line][3]
                fornecedor = dictLine[line][4]
                cnpj = dictLine[line][5]
                dataVencimento = dictLine[line][6]
                
                row.append(idHeader)# '*Invoice ID'
                row.append("1")# 'Line Number'
                row.append("ITEM")# '*Line Type'
                row.append(saldo)# '*Amount'
                row.append("")# 'Invoice Quantity'
                row.append("")# 'Unit Price'
                row.append("")# 'UOM'
                row.append("")# 'Description'
                row.append("")# 'PO Number'
                row.append("")# 'PO Line Number'
                row.append("")# 'PO Schedule Number'
                row.append("")# 'PO Distribution Number'
                row.append("")# 'Item Description'
                row.append("")# 'PO Release Number'
                row.append("")# 'Purchasing Category'
                row.append("")# 'Receipt Number'
                row.append("")# 'Receipt Line Number'
                row.append("")# 'Consumption Advice Number'
                row.append("")# 'Consumption Advice Line Number'
                row.append("")# 'Packing Slip'
                row.append("N")# 'Final Match'
                row.append("1001.000.000.11470001.0000.0000.0.0")# 'Distribution Combination'
                row.append("")# 'Distribution Set'
                row.append("30/04/2022")# 'Accounting Date'
                row.append("")# 'Overlay Account Segment'
                row.append("")# 'Overlay Primary Balancing Segment'
                row.append("")# 'Overlay Cost Center Segment'
                row.append("")# 'Tax Classification Code'
                row.append("LOC_MATRIZ")# 'Ship-to Location'
                row.append("")# 'Ship-from Location'
                row.append("")# 'Location of Final Discharge'
                row.append("")# 'Transaction Business Category'
                row.append("")# 'Product Fiscal Classification'
                row.append("")# 'Intended Use'
                row.append("")# 'User-Defined Fiscal Classification'
                row.append("")# 'Product Type'
                row.append("")# 'Assessable Value'
                row.append("")# 'Product Category'
                row.append("")# 'Tax Control Amount'
                row.append("")# 'Tax Regime Code'
                row.append("")# 'Tax'
                row.append("")# 'Tax Status Code'
                row.append("")# 'Tax Jurisdiction Code'
                row.append("")# 'Tax Rate Code'
                row.append("")# 'Tax Rate'
                row.append("")# 'Withholding Tax Group'
                row.append("")# 'Income Tax Type'
                row.append("")# 'Income Tax Region'
                row.append("N")# 'Prorate Across All Item Lines'
                row.append("1")# 'Line Group Number'
                row.append("")# 'Cost Factor Name'
                row.append("")# 'Statistical Quantity'
                row.append("N")# 'Track as Asset'
                row.append("")# 'Asset Book Type Code'
                row.append("")# 'Asset Category ID'
                row.append("")# 'Serial Number'
                row.append("")# 'Manufacturer'
                row.append("")# 'Model Number'
                row.append("")# 'Warranty Number'
                row.append("N")# 'Price Correction Line'
                row.append("")# 'Price Correction Invoice Number'
                row.append("")# 'Price Correction Invoice Line Number'
                row.append("")# 'Requester First Name'
                row.append("")# 'Requester Last Name'
                row.append("")# 'Requester Employee Number'
                row.append("")# 'Attribute Category'
                row.append("")# 'Attribute 1'
                row.append("")# 'Attribute 2'
                row.append("")# 'Attribute 3'
                row.append("")# 'Attribute 4'
                row.append("")# 'Attribute 5'
                row.append("")# 'Attribute 6'
                row.append("")# 'Attribute 7'
                row.append("")# 'Attribute 8'
                row.append("")# 'Attribute 9'
                row.append("")# 'Attribute 10'
                row.append("")# 'Attribute 11'
                row.append("")# 'Attribute 12'
                row.append("")# 'Attribute 13'
                row.append("")# 'Attribute 14'
                row.append("")# 'Attribute 15'
                row.append("")# 'Attribute Number 1'
                row.append("")# 'Attribute Number 2'
                row.append("")# 'Attribute Number 3'
                row.append("")# 'Attribute Number 4'
                row.append("")# 'Attribute Number 5'
                row.append("")# 'Attribute Date 1'
                row.append("")# 'Attribute Date 2'
                row.append("")# 'Attribute Date 3'
                row.append("")# 'Attribute Date 4'
                row.append("")# 'Attribute Date 5'
                row.append("")# 'Global Attribute Category'
                row.append("")# 'Global Attribute 1'
                row.append("")# 'Global Attribute 2'
                row.append("")# 'Global Attribute 3'
                row.append("")# 'Global Attribute 4'
                row.append("")# 'Global Attribute 5'
                row.append("")# 'Global Attribute 6'
                row.append("")# 'Global Attribute 7'
                row.append("")# 'Global Attribute 8'
                row.append("")# 'Global Attribute 9'
                row.append("")# 'Global Attribute 10'
                row.append("")# 'Global Attribute 11'
                row.append("")# 'Global Attribute 12'
                row.append("")# 'Global Attribute 13'
                row.append("")# 'Global Attribute 14'
                row.append("")# 'Global Attribute 15'
                row.append("")# 'Global Attribute 16'
                row.append("")# 'Global Attribute 17'
                row.append("")# 'Global Attribute 18'
                row.append("")# 'Global Attribute 19'
                row.append("")# 'Global Attribute 20'
                row.append("")# 'Global Attribute Number 1'
                row.append("")# 'Global Attribute Number 2'
                row.append("")# 'Global Attribute Number 3'
                row.append("")# 'Global Attribute Number 4'
                row.append("")# 'Global Attribute Number 5'
                row.append("")# 'Global Attribute Date 1'
                row.append("")# 'Global Attribute Date 2'
                row.append("")# 'Global Attribute Date 3'
                row.append("")# 'Global Attribute Date 4'
                row.append("")# 'Global Attribute Date 5'
                row.append("")# 'Project ID'
                row.append("")# 'Task ID'
                row.append("")# 'Expenditure Type ID'
                row.append("")# 'Expenditure Item Date'
                row.append("")# 'Expenditure Organization ID'
                row.append("")# 'Project Number'
                row.append("")# 'Task Number'
                row.append("")# 'Expenditure Type'
                row.append("")# 'Expenditure Organization'
                row.append("")# 'Funding Source Id'
                row.append("")# 'PJC Reserved Attribute 2'
                row.append("")# 'PJC Reserved Attribute 3'
                row.append("")# 'PJC Reserved Attribute 4'
                row.append("")# 'PJC Reserved Attribute 5'
                row.append("")# 'PJC Reserved Attribute 6'
                row.append("")# 'PJC Reserved Attribute 7'
                row.append("")# 'PJC Reserved Attribute 8'
                row.append("")# 'PJC Reserved Attribute 9'
                row.append("")# 'PJC Reserved Attribute 10'
                row.append("")# 'PJC User Defined Attribute 1'
                row.append("")# 'PJC User Defined Attribute 2'
                row.append("")# 'PJC User Defined Attribute 3'
                row.append("")# 'PJC User Defined Attribute 4'
                row.append("")# 'PJC User Defined Attribute 5'
                row.append("")# 'PJC User Defined Attribute 6'
                row.append("")# 'PJC User Defined Attribute 7'
                row.append("")# 'PJC User Defined Attribute 8'
                row.append("")# 'PJC User Defined Attribute 9'
                row.append("")# 'PJC User Defined Attribute 10'
                row.append("")# 'Fiscal Charge Type'
                row.append("")# 'Multiperiod Accounting Start Date'
                row.append("")# 'Multiperiod Accounting End Date'
                row.append("")# 'Multiperiod Accounting Accrual Account'  
                row.append("")# 'Project Name'	
                row.append("")# 'Task Name'

                ws.append(row)

        wb.save("Plan_Contas_AP_ADIANTAMENTO_LINE_TESTE.xlsx")

line()